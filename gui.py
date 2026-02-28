"""
gui.py — PyQt6 single-window dashboard for the Drowsiness Monitoring System.
Displays live camera feed, real-time metrics, and a fatigue score progress bar.
Uses qtawesome for Font Awesome icons and supports Excel export.
"""

import time
import os
from datetime import datetime

import cv2
import numpy as np
from openpyxl import Workbook
from utils import resource_path
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QLabel, QVBoxLayout, QHBoxLayout,
    QProgressBar, QFrame, QSizePolicy, QPushButton, QFileDialog,
    QMessageBox, QGraphicsDropShadowEffect,
)
from PyQt6.QtCore import QTimer, Qt, QSize
from PyQt6.QtGui import QImage, QPixmap, QFont, QColor, QLinearGradient, QPainter
import qtawesome as qta

import config
from detector import Detector
from fatigue_engine import FatigueEngine, FatigueState
from alarm import AlarmManager
from logger import SessionLogger


# ─── Colour Palette ──────────────────────────────────────────────────────────

DARK_BG       = "#090c10"
PANEL_BG      = "#0d1117"
CARD_BG       = "#161b22"
CARD_BORDER   = "#21262d"
TEXT_PRIMARY   = "#e6edf3"
TEXT_SECONDARY = "#8b949e"
TEXT_DIM       = "#484f58"
ACCENT_CYAN    = "#58a6ff"
ACCENT_GREEN   = "#3fb950"
ACCENT_PURPLE  = "#bc8cff"
BORDER_NORMAL     = "#238636"
BORDER_WARNING1   = "#d29922"
BORDER_WARNING2   = "#db6d28"
BORDER_CRITICAL   = "#f85149"
BORDER_CALIBRATING = "#58a6ff"

GLOBAL_STYLESHEET = f"""
    QMainWindow {{
        background-color: {DARK_BG};
    }}
    QWidget {{
        background-color: {DARK_BG};
        color: {TEXT_PRIMARY};
    }}
    QLabel {{
        color: {TEXT_PRIMARY};
        background-color: transparent;
    }}
    QScrollBar:vertical {{
        background: {DARK_BG};
        width: 6px;
    }}
    QScrollBar::handle:vertical {{
        background: {CARD_BORDER};
        border-radius: 3px;
    }}
"""

EXPORT_BTN_STYLE = f"""
    QPushButton {{
        background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
            stop:0 #238636, stop:1 #2ea043);
        color: white;
        border: none;
        border-radius: 6px;
        padding: 8px 14px;
        font-weight: 600;
        font-size: 11px;
    }}
    QPushButton:hover {{
        background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
            stop:0 #2ea043, stop:1 #3fb950);
    }}
    QPushButton:pressed {{
        background: #238636;
    }}
"""


class StatCard(QFrame):
    """A compact card showing an icon, label, and value."""

    def __init__(self, title: str, icon_name: str = "", parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            StatCard {{
                background-color: {CARD_BG};
                border: 1px solid {CARD_BORDER};
                border-radius: 8px;
            }}
        """)
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setFixedHeight(48)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 4, 10, 4)
        layout.setSpacing(8)

        # Icon
        if icon_name:
            icon_label = QLabel()
            icon_pixmap = qta.icon(icon_name, color=TEXT_DIM).pixmap(14, 14)
            icon_label.setPixmap(icon_pixmap)
            icon_label.setFixedSize(16, 16)
            layout.addWidget(icon_label)

        # Title
        title_label = QLabel(title)
        title_label.setFont(QFont("Segoe UI", 9))
        title_label.setStyleSheet(f"color: {TEXT_SECONDARY};")
        layout.addWidget(title_label)

        layout.addStretch()

        # Value (right-aligned)
        self.value_label = QLabel("—")
        self.value_label.setFont(QFont("Segoe UI Semibold", 13))
        self.value_label.setStyleSheet(f"color: {ACCENT_CYAN};")
        self.value_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        layout.addWidget(self.value_label)

    def set_value(self, text: str, color: str = ACCENT_CYAN):
        self.value_label.setText(text)
        self.value_label.setStyleSheet(f"color: {color};")


class CalibrationOverlay(QWidget):
    """A styled circular progress overlay shown during calibration."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(220, 220)
        self._progress = 0.0  # 0.0 to 1.0
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)

    def set_progress(self, value: float):
        self._progress = max(0.0, min(1.0, value))
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        size = min(self.width(), self.height())
        margin = 15
        rect_size = size - 2 * margin

        # Background circle
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(0, 0, 0, 160))
        painter.drawEllipse(margin, margin, rect_size, rect_size)

        # Progress arc
        pen_width = 6
        from PyQt6.QtCore import QRectF
        from PyQt6.QtGui import QPen
        arc_rect = QRectF(margin + pen_width / 2, margin + pen_width / 2,
                          rect_size - pen_width, rect_size - pen_width)

        # Track
        track_pen = QPen(QColor(CARD_BORDER), pen_width)
        track_pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        painter.setPen(track_pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawEllipse(arc_rect)

        # Active arc
        progress_pen = QPen(QColor(ACCENT_CYAN), pen_width)
        progress_pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        painter.setPen(progress_pen)
        span = int(-self._progress * 360 * 16)  # Qt uses 1/16th degrees
        painter.drawArc(arc_rect, 90 * 16, span)

        # Percentage text
        painter.setPen(QColor(TEXT_PRIMARY))
        painter.setFont(QFont("Segoe UI Black", 28))
        pct_text = f"{int(self._progress * 100)}%"
        painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, pct_text)

        # "Calibrating" label below percentage
        painter.setFont(QFont("Segoe UI", 9))
        painter.setPen(QColor(TEXT_SECONDARY))
        sub_rect = self.rect().adjusted(0, 40, 0, 0)
        painter.drawText(sub_rect, Qt.AlignmentFlag.AlignCenter, "CALIBRATING")

        painter.end()


class MainWindow(QMainWindow):
    """Main application window — drowsiness monitoring dashboard."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle(config.WINDOW_TITLE)
        self.setMinimumSize(config.WINDOW_MIN_WIDTH, config.WINDOW_MIN_HEIGHT)
        self.setStyleSheet(GLOBAL_STYLESHEET)

        # ── Subsystems ──
        self.detector = Detector()
        self.engine = FatigueEngine()
        self.alarm_manager = AlarmManager()
        self.logger = SessionLogger()

        # ── FPS tracking ──
        self._frame_times = []
        self._fps = 0.0

        # ── Alarm event history (for Excel export) ──
        self._alarm_events = []
        self._metrics_history = []
        self._last_alarm_level = config.STATUS_NORMAL
        self._last_snapshot_time = 0.0

        # ── Build UI ──
        self._build_ui()

        # ── Start timer ──
        self.timer = QTimer(self)
        self.timer.timeout.connect(self._process_frame)
        self.timer.start(config.GUI_TIMER_INTERVAL_MS)

    # ──────────────────────────────────────────────────────────────────────────
    # UI Construction
    # ──────────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)

        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(8)

        # ── Header bar ──
        header = QHBoxLayout()
        header.setSpacing(16)

        logo_icon = QLabel()
        pixmap = QPixmap(resource_path("assets/logo.png"))
        logo_icon.setPixmap(pixmap.scaledToHeight(40, Qt.TransformationMode.SmoothTransformation))
        logo_icon.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        header.addWidget(logo_icon)

        header.addStretch()

        # Session timer
        self._session_start = time.time()
        self.session_label = QLabel()
        self.session_label.setFont(QFont("Segoe UI", 9))
        self.session_label.setStyleSheet(f"color: {TEXT_SECONDARY};")
        header.addWidget(self.session_label)

        # Fullscreen toggle button
        self.fullscreen_btn = QPushButton()
        self.fullscreen_btn.setIcon(qta.icon("fa5s.expand", color=TEXT_SECONDARY))
        self.fullscreen_btn.setFixedSize(30, 30)
        self.fullscreen_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.fullscreen_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {CARD_BG};
                border: 1px solid {CARD_BORDER};
                border-radius: 6px;
            }}
            QPushButton:hover {{
                background-color: #30363d;
            }}
        """)
        self.fullscreen_btn.setToolTip("Toggle Fullscreen (F11)")
        self.fullscreen_btn.clicked.connect(self._toggle_fullscreen)
        header.addWidget(self.fullscreen_btn)

        main_layout.addLayout(header)

        # ── Separator ──
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet(f"background-color: {CARD_BORDER}; max-height: 1px;")
        main_layout.addWidget(sep)

        # ── Top row: camera (left) + stats (right) ──
        top_layout = QHBoxLayout()
        top_layout.setSpacing(8)

        # === Left panel — camera feed ===
        self.camera_frame = QFrame()
        self.camera_frame.setStyleSheet(f"""
            QFrame {{
                background-color: #000;
                border: 3px solid {BORDER_CALIBRATING};
                border-radius: 12px;
            }}
        """)
        camera_inner = QVBoxLayout(self.camera_frame)
        camera_inner.setContentsMargins(0, 0, 0, 0)
        camera_inner.setSpacing(0)

        # Video label
        self.video_label = QLabel()
        self.video_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.video_label.setMinimumSize(560, 420)
        self.video_label.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding
        )
        camera_inner.addWidget(self.video_label)

        # Status bar at bottom of camera frame
        status_bar = QFrame()
        status_bar.setFixedHeight(42)
        status_bar.setStyleSheet(f"""
            QFrame {{
                background-color: rgba(13, 17, 23, 220);
                border-bottom-left-radius: 10px;
                border-bottom-right-radius: 10px;
                border: none;
            }}
        """)
        status_layout = QHBoxLayout(status_bar)
        status_layout.setContentsMargins(16, 0, 16, 0)

        self.status_icon = QLabel()
        self.status_icon.setFixedSize(18, 18)
        status_layout.addWidget(self.status_icon)

        self.status_label = QLabel(config.STATUS_CALIBRATING)
        self.status_label.setFont(QFont("Segoe UI Black", 13))
        self.status_label.setStyleSheet(f"color: {ACCENT_CYAN};")
        status_layout.addWidget(self.status_label)

        status_layout.addStretch()

        # Live EAR in camera status bar
        self.live_ear_label = QLabel("EAR: —")
        self.live_ear_label.setFont(QFont("Segoe UI", 10))
        self.live_ear_label.setStyleSheet(f"color: {TEXT_SECONDARY};")
        status_layout.addWidget(self.live_ear_label)

        camera_inner.addWidget(status_bar)

        # Calibration overlay (positioned over the video label)
        self.calibration_overlay = CalibrationOverlay(self.video_label)
        self.calibration_overlay.setVisible(True)

        top_layout.addWidget(self.camera_frame, stretch=75)

        # === Right panel — stats ===
        stats_panel = QFrame()
        stats_panel.setStyleSheet(f"""
            QFrame {{
                background-color: {PANEL_BG};
                border: 1px solid {CARD_BORDER};
                border-radius: 10px;
            }}
        """)
        stats_layout = QVBoxLayout(stats_panel)
        stats_layout.setContentsMargins(8, 8, 8, 8)
        stats_layout.setSpacing(4)

        # Panel header
        panel_header = QHBoxLayout()
        panel_header.setSpacing(6)
        chart_lbl = QLabel()
        chart_lbl.setPixmap(qta.icon("fa5s.chart-line", color=ACCENT_PURPLE).pixmap(14, 14))
        chart_lbl.setFixedSize(16, 16)
        panel_header.addWidget(chart_lbl)
        panel_title = QLabel("Live Metrics")
        panel_title.setFont(QFont("Segoe UI Semibold", 10))
        panel_title.setStyleSheet(f"color: {TEXT_PRIMARY};")
        panel_header.addWidget(panel_title)
        panel_header.addStretch()
        stats_layout.addLayout(panel_header)

        # Stat cards
        self.card_ear       = StatCard("Current EAR",       "fa5s.eye")
        self.card_baseline  = StatCard("Baseline EAR",      "fa5s.crosshairs")
        self.card_threshold = StatCard("EAR Threshold",     "fa5s.sliders-h")
        self.card_blink_dur = StatCard("Blink Duration",    "fa5s.clock")
        self.card_blink_cnt = StatCard("Blinks (60s)",      "fa5s.hand-paper")
        self.card_consec    = StatCard("Low Frames",        "fa5s.layer-group")
        self.card_fatigue   = StatCard("Fatigue Score",     "fa5s.brain")
        self.card_microsleep = StatCard("Micro-sleeps",     "fa5s.moon")
        self.card_fps       = StatCard("FPS",               "fa5s.tachometer-alt")

        for card in [
            self.card_ear, self.card_baseline, self.card_threshold,
            self.card_blink_dur, self.card_blink_cnt, self.card_consec,
            self.card_fatigue, self.card_microsleep, self.card_fps,
        ]:
            stats_layout.addWidget(card)

        stats_layout.addStretch()

        # Export button
        self.export_btn = QPushButton("  Export Report")
        self.export_btn.setIcon(qta.icon("fa5s.file-excel", color="white"))
        self.export_btn.setStyleSheet(EXPORT_BTN_STYLE)
        self.export_btn.setFixedHeight(36)
        self.export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.export_btn.clicked.connect(self._export_to_excel)
        stats_layout.addWidget(self.export_btn)

        top_layout.addWidget(stats_panel, stretch=25)
        main_layout.addLayout(top_layout, stretch=1)

        # ── Bottom: fatigue progress bar ──
        bottom_frame = QFrame()
        bottom_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {PANEL_BG};
                border: 1px solid {CARD_BORDER};
                border-radius: 8px;
            }}
        """)
        bottom_layout = QHBoxLayout(bottom_frame)
        bottom_layout.setContentsMargins(14, 8, 14, 8)
        bottom_layout.setSpacing(10)

        bolt_icon = QLabel()
        bolt_icon.setPixmap(qta.icon("fa5s.heartbeat", color="#f85149").pixmap(16, 16))
        bolt_icon.setFixedSize(18, 18)
        bottom_layout.addWidget(bolt_icon)

        bar_label = QLabel("Fatigue")
        bar_label.setFont(QFont("Segoe UI Semibold", 10))
        bar_label.setStyleSheet(f"color: {TEXT_PRIMARY};")
        bar_label.setFixedWidth(55)
        bottom_layout.addWidget(bar_label)

        self.fatigue_bar = QProgressBar()
        self.fatigue_bar.setRange(0, 100)
        self.fatigue_bar.setValue(0)
        self.fatigue_bar.setFixedHeight(22)
        self.fatigue_bar.setTextVisible(True)
        self.fatigue_bar.setFormat("%v")
        self.fatigue_bar.setFont(QFont("Segoe UI Semibold", 9))
        self._set_bar_color(BORDER_NORMAL)
        bottom_layout.addWidget(self.fatigue_bar)

        # Tier markers
        for level, threshold in [("W", 25), ("OFF", 40), ("CR", 60)]:
            marker = QLabel(f"{level}:{threshold}")
            marker.setFont(QFont("Segoe UI", 8))
            marker.setStyleSheet(f"color: {TEXT_DIM};")
            marker.setFixedWidth(40)
            bottom_layout.addWidget(marker)

        main_layout.addWidget(bottom_frame)

    # ──────────────────────────────────────────────────────────────────────────
    # Main Loop
    # ──────────────────────────────────────────────────────────────────────────

    def _process_frame(self):
        now = time.time()

        detection = self.detector.get_frame()
        if detection.frame is None:
            return

        state = self.engine.update(detection.ear_avg, detection.face_detected)

        # Alarm
        if not state.is_calibrating:
            self.alarm_manager.update(state.alarm_level)

        # Track alarm events
        self._track_alarm_events(state, now)

        # Periodic metric snapshot
        if now - self._last_snapshot_time >= 2.0:
            self._metrics_history.append({
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "ear": round(state.current_ear, 4),
                "baseline_ear": round(state.baseline_ear, 4),
                "ear_threshold": round(state.ear_threshold, 4),
                "fatigue_score": round(state.fatigue_score, 1),
                "blink_count_60s": state.blink_count_last_60s,
                "microsleep_count": state.microsleep_count,
                "alarm_level": state.alarm_level,
                "event": state.last_event,
            })
            self._last_snapshot_time = now

        # CSV log
        if state.last_event or state.is_calibrating:
            self.logger.log_event(
                ear=state.current_ear,
                fatigue_score=state.fatigue_score,
                event_type=state.last_event if state.last_event else (
                    "calibrating" if state.is_calibrating else ""
                ),
                alarm_level=state.alarm_level,
                blink_duration=state.blink_duration,
                microsleep_detected=(state.last_event == "microsleep"),
            )

        # FPS
        self._frame_times.append(now)
        if len(self._frame_times) > 30:
            self._frame_times = self._frame_times[-30:]
        if len(self._frame_times) >= 2:
            elapsed = self._frame_times[-1] - self._frame_times[0]
            if elapsed > 0:
                self._fps = (len(self._frame_times) - 1) / elapsed

        # Update GUI
        self._update_video(detection, state)
        self._update_calibration_overlay(state)
        self._update_stats(state)
        self._update_status(state, detection.face_detected)
        self._update_fatigue_bar(state)
        self._update_session_timer()

    def _track_alarm_events(self, state: FatigueState, now: float):
        current_level = state.alarm_level
        if current_level != self._last_alarm_level:
            if current_level in (config.STATUS_WARNING_1,
                                 config.STATUS_CRITICAL):
                self._alarm_events.append({
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                    "alarm_level": current_level,
                    "fatigue_score": round(state.fatigue_score, 1),
                    "ear_at_trigger": round(state.current_ear, 4),
                    "microsleep_count": state.microsleep_count,
                    "blink_count_60s": state.blink_count_last_60s,
                    "trigger_event": state.last_event if state.last_event else "score_threshold",
                })
            self._last_alarm_level = current_level

    # ──────────────────────────────────────────────────────────────────────────
    # Excel Export
    # ──────────────────────────────────────────────────────────────────────────

    def _export_to_excel(self):
        default_name = f"drowsiness_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        default_path = os.path.join(os.path.expanduser("~"), "Desktop", default_name)

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Report to Excel", default_path, "Excel Files (*.xlsx)",
        )
        if not file_path:
            return

        try:
            wb = Workbook()

            # Sheet 1: Metrics
            ws1 = wb.active
            ws1.title = "Real-Time Metrics"
            headers = ["Timestamp", "EAR", "Baseline EAR", "EAR Threshold",
                        "Fatigue Score", "Blinks (60s)", "Microsleep Count",
                        "Alarm Level", "Event"]
            ws1.append(headers)
            for row in self._metrics_history:
                ws1.append([row["timestamp"], row["ear"], row["baseline_ear"],
                            row["ear_threshold"], row["fatigue_score"],
                            row["blink_count_60s"], row["microsleep_count"],
                            row["alarm_level"], row["event"]])

            from openpyxl.styles import Font as XlFont, PatternFill
            hf = XlFont(bold=True, color="FFFFFF")
            hfill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            for cell in ws1[1]:
                cell.font = hf
                cell.fill = hfill
            for col in ws1.columns:
                mx = max(len(str(c.value or "")) for c in col)
                ws1.column_dimensions[col[0].column_letter].width = mx + 3

            # Sheet 2: Alarm Events
            ws2 = wb.create_sheet("Alarm Events")
            alarm_headers = ["Timestamp", "Alarm Level", "Fatigue Score",
                             "EAR at Trigger", "Microsleep Count",
                             "Blinks (60s)", "Trigger Event"]
            ws2.append(alarm_headers)
            for row in self._alarm_events:
                ws2.append([row["timestamp"], row["alarm_level"],
                            row["fatigue_score"], row["ear_at_trigger"],
                            row["microsleep_count"], row["blink_count_60s"],
                            row["trigger_event"]])
            for cell in ws2[1]:
                cell.font = hf
                cell.fill = hfill
            for col in ws2.columns:
                mx = max(len(str(c.value or "")) for c in col)
                ws2.column_dimensions[col[0].column_letter].width = mx + 3

            # Sheet 3: Summary
            ws3 = wb.create_sheet("Session Summary")
            ws3.append(["Metric", "Value"])
            elapsed_min = (time.time() - self._session_start) / 60
            summary = [
                ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("Session Duration", f"{elapsed_min:.1f} minutes"),
                ("Metric Snapshots", len(self._metrics_history)),
                ("Total Alarm Events", len(self._alarm_events)),
                ("Warning Events", sum(1 for e in self._alarm_events if e["alarm_level"] == config.STATUS_WARNING_1)),
                ("Critical Events", sum(1 for e in self._alarm_events if e["alarm_level"] == config.STATUS_CRITICAL)),
            ]
            for m, v in summary:
                ws3.append([m, v])
            for cell in ws3[1]:
                cell.font = hf
                cell.fill = hfill
            ws3.column_dimensions["A"].width = 25
            ws3.column_dimensions["B"].width = 30

            wb.save(file_path)
            QMessageBox.information(self, "Export Successful",
                f"Report saved to:\n{file_path}\n\n"
                f"• {len(self._metrics_history)} metric snapshots\n"
                f"• {len(self._alarm_events)} alarm events")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not save:\n{str(e)}")

    # ──────────────────────────────────────────────────────────────────────────
    # GUI Updates
    # ──────────────────────────────────────────────────────────────────────────

    def _update_video(self, detection, state: FatigueState):
        frame = (detection.annotated_frame
                 if detection.annotated_frame is not None
                 else detection.frame)
        if frame is None:
            return

        if not detection.face_detected and not state.is_calibrating:
            cv2.putText(frame, "Align face with camera", (30, 50),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.0, config.COLOR_YELLOW, 2)

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb.shape
        q_img = QImage(rgb.data, w, h, ch * w, QImage.Format.Format_RGB888)
        pixmap = QPixmap.fromImage(q_img)
        scaled = pixmap.scaled(
            self.video_label.size(),
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
        self.video_label.setPixmap(scaled)

    def _update_calibration_overlay(self, state: FatigueState):
        if state.is_calibrating:
            self.calibration_overlay.setVisible(True)
            self.calibration_overlay.set_progress(state.calibration_progress)
            # Center the overlay on the video label
            vw, vh = self.video_label.width(), self.video_label.height()
            ow, oh = self.calibration_overlay.width(), self.calibration_overlay.height()
            self.calibration_overlay.move((vw - ow) // 2, (vh - oh) // 2)
        else:
            self.calibration_overlay.setVisible(False)

    def _update_stats(self, state: FatigueState):
        self.card_ear.set_value(f"{state.current_ear:.3f}")
        self.card_baseline.set_value(
            f"{state.baseline_ear:.3f}" if state.baseline_ear > 0 else "—"
        )
        self.card_threshold.set_value(f"{state.ear_threshold:.3f}")
        self.card_blink_dur.set_value(f"{state.blink_duration:.2f}s")
        self.card_blink_cnt.set_value(str(state.blink_count_last_60s))
        self.card_consec.set_value(str(state.consecutive_low_frames))

        score = state.fatigue_score
        if score >= 60:
            sc = BORDER_CRITICAL
        elif score >= 40:
            sc = BORDER_WARNING1
        else:
            sc = BORDER_NORMAL
        self.card_fatigue.set_value(f"{score:.0f}", color=sc)
        self.card_microsleep.set_value(
            str(state.microsleep_count),
            color=BORDER_CRITICAL if state.microsleep_count > 0 else ACCENT_CYAN,
        )
        self.card_fps.set_value(f"{self._fps:.1f}")

        # Live EAR in camera bar
        self.live_ear_label.setText(f"EAR: {state.current_ear:.3f}")

    def _update_status(self, state: FatigueState, face_detected: bool):
        if state.is_calibrating:
            status_text = config.STATUS_CALIBRATING
            border = BORDER_CALIBRATING
            text_color = ACCENT_CYAN
            icon_name = "fa5s.sync-alt"
        elif not face_detected:
            status_text = config.STATUS_NO_FACE
            border = BORDER_WARNING1
            text_color = BORDER_WARNING1
            icon_name = "fa5s.user-slash"
        elif state.alarm_level == config.STATUS_CRITICAL:
            status_text = config.STATUS_CRITICAL
            border = BORDER_CRITICAL
            text_color = BORDER_CRITICAL
            icon_name = "fa5s.exclamation-triangle"
        elif state.alarm_level == config.STATUS_WARNING_1:
            status_text = config.STATUS_WARNING_1
            border = BORDER_WARNING1
            text_color = BORDER_WARNING1
            icon_name = "fa5s.bell"
        else:
            status_text = config.STATUS_NORMAL
            border = BORDER_NORMAL
            text_color = BORDER_NORMAL
            icon_name = "fa5s.check-circle"

        self.status_label.setText(status_text)
        self.status_label.setStyleSheet(f"color: {text_color};")
        self.status_icon.setPixmap(qta.icon(icon_name, color=text_color).pixmap(16, 16))

        self.camera_frame.setStyleSheet(f"""
            QFrame {{
                background-color: #000;
                border: 3px solid {border};
                border-radius: 12px;
            }}
        """)

    def _update_fatigue_bar(self, state: FatigueState):
        score = int(state.fatigue_score)
        self.fatigue_bar.setValue(score)
        if score >= 60:
            self._set_bar_color(BORDER_CRITICAL)
        elif score >= 40:
            self._set_bar_color(BORDER_WARNING1)
        else:
            self._set_bar_color(BORDER_NORMAL)

    def _set_bar_color(self, color: str):
        self.fatigue_bar.setStyleSheet(f"""
            QProgressBar {{
                background-color: {CARD_BG};
                border: 1px solid {CARD_BORDER};
                border-radius: 5px;
                text-align: center;
                color: {TEXT_PRIMARY};
            }}
            QProgressBar::chunk {{
                background-color: {color};
                border-radius: 4px;
            }}
        """)

    def _update_session_timer(self):
        elapsed = int(time.time() - self._session_start)
        m, s = divmod(elapsed, 60)
        h, m = divmod(m, 60)
        self.session_label.setText(f"Session: {h:02d}:{m:02d}:{s:02d}")

    def _toggle_fullscreen(self):
        """Toggle between windowed and fullscreen mode."""
        if self.isFullScreen():
            self.showNormal()
            self.fullscreen_btn.setIcon(qta.icon("fa5s.expand", color=TEXT_SECONDARY))
            self.fullscreen_btn.setToolTip("Enter Fullscreen (F11)")
        else:
            self.showFullScreen()
            self.fullscreen_btn.setIcon(qta.icon("fa5s.compress", color=TEXT_SECONDARY))
            self.fullscreen_btn.setToolTip("Exit Fullscreen (F11)")

    def keyPressEvent(self, event):
        """Handle keyboard shortcuts."""
        if event.key() == Qt.Key.Key_F11:
            self._toggle_fullscreen()
        elif event.key() == Qt.Key.Key_Escape and self.isFullScreen():
            self._toggle_fullscreen()
        else:
            super().keyPressEvent(event)

    # ──────────────────────────────────────────────────────────────────────────────
    # Cleanup
    # ──────────────────────────────────────────────────────────────────────────────

    def closeEvent(self, event):
        self.timer.stop()
        self.detector.release()
        self.alarm_manager.cleanup()
        self.logger.close()
        event.accept()
